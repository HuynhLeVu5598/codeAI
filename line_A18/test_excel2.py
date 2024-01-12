# # import openpyxl
   
# # # Define variable to load the dataframe
# # wb = openpyxl.Workbook()
   
# # # Define active sheet
# # sheet = wb.active
  
# # # Create List for store data
# # data =[('ID', 'Name', 'Email'),
# #        (1, 'Hardik Savani', 'hardik@gmail.com'),
# #        (2, 'Vimal Kashiyani', 'vimal@gmail.com'),
# #        (3, 'Harshad Pathak', 'harshad@gmail.com')]
  
# # # Adding Data to Sheet
# # for item in data :
# #      sheet.append(item)
  
# # # Save File
# # wb.save("excel/demo.xlsx")


import openpyxl
from openpyxl.styles import Alignment, Font
from datetime import date
import datetime
# Define variable to load the dataframe
wb = openpyxl.Workbook()

#Ngay = wb.create_sheet("Ngay")
  
today = date.today()
d1 = today.strftime("%d_%m_%Y")

now = datetime.datetime.now()
t1 = now.strftime("%H-%M-%S")


# # Ngay.merge_cells('A1:L1')
# # Ngay.merge_cells('A2:A3')
# # Ngay.merge_cells('B2:B3')
# # Ngay.merge_cells('C2:L2')
# # #Ngay.unmerge_cells('A2:D2')
# # Ngay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
# # Ngay['A1'].alignment = Alignment(horizontal='center')
# # Ngay['A1'].font = Font(name= 'Calibri', size=20)

# # Ngay['A2'] = 'Ngày sản xuất'
# # Ngay['A2'].alignment = Alignment(horizontal='center')
# # Ngay['A2'].font = Font(name= 'Calibri', size=12)
# # Ngay['B2'] = 'Giờ lưu dữ liệu'
# # Ngay['B2'].alignment = Alignment(horizontal='center')
# # Ngay['B2'].font = Font(name= 'Calibri', size=12)
# # Ngay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
# # Ngay['C2'].alignment = Alignment(horizontal='center')
# # Ngay['C2'].font = Font(name= 'Calibri', size=12)
# # Ngay['C3'] = 'Tổng số lượng sản xuất'
# # Ngay['D3'] = 'Cuộn cảm'
# # Ngay['E3'] = 'Cacbon tay chổi'
# # Ngay['F3'] = 'Hàn chổi'
# # Ngay['G3'] = 'Hàn chấu '
# # Ngay['H3'] = 'Đế vỏ nhỏ'
# # Ngay['I3'] = 'Tụ điện'
# # Ngay['J3'] = 'Cong chấu điện'
# # Ngay['K3'] = 'Bụi chì'
# # Ngay['L3'] = 'PP khác (nhiều hạng mục)'

# # for i in range(67,77):
# #     Ngay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
# #     Ngay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 


# Dem = wb.create_sheet("All")
  
# Dem.merge_cells('A1:L1')
# Dem.merge_cells('A2:A3')
# Dem.merge_cells('B2:B3')
# Dem.merge_cells('C2:L2')
# #Dem.unmerge_cells('A2:D2')
# Dem['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
# Dem['A1'].alignment = Alignment(horizontal='center')
# Dem['A1'].font = Font(name= 'Calibri', size=20)

# Dem['A2'] = 'Ngày sản xuất'
# Dem['A2'].alignment = Alignment(horizontal='center')
# Dem['A2'].font = Font(name= 'Calibri', size=12)
# Dem['B2'] = 'Giờ lưu dữ liệu'
# Dem['B2'].alignment = Alignment(horizontal='center')
# Dem['B2'].font = Font(name= 'Calibri', size=12)
# Dem['C2'] = 'HẠNG MỤC PHẾ PHẨM'
# Dem['C2'].alignment = Alignment(horizontal='center')
# Dem['C2'].font = Font(name= 'Calibri', size=12)
# Dem['C3'] = 'Tổng số lượng sản xuất'
# Dem['D3'] = 'Cuộn cảm'
# Dem['E3'] = 'Cacbon tay chổi'
# Dem['F3'] = 'Hàn chổi'
# Dem['G3'] = 'Hàn chấu '
# Dem['H3'] = 'Đế vỏ nhỏ'
# Dem['I3'] = 'Tụ điện'
# Dem['J3'] = 'Cong chấu điện'
# Dem['K3'] = 'Bụi chì'
# Dem['L3'] = 'PP khác (nhiều hạng mục)'

# for i in range(67,77):
#     Dem[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
#     Dem[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

   
# # Remove default Sheet
# wb.remove(wb['Sheet'])
  
# # Iterate the loop to read the cell values
# wb.save(f"excel/All.xlsx")



today = date.today()
mydate = today.strftime("%Y_%m_%d")
wb = openpyxl.Workbook()

HomNay = wb.create_sheet("Data")


HomNay.merge_cells('A1:L1')
HomNay.merge_cells('A2:A3')
HomNay.merge_cells('B2:B3')
HomNay.merge_cells('C2:L2')
#HomNay.unmerge_cells('A2:D2')
HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
HomNay['A1'].alignment = Alignment(horizontal='center')
HomNay['A1'].font = Font(name= 'Calibri', size=20)

HomNay['A2'] = 'Ngày sản xuất'
HomNay['A2'].alignment = Alignment(horizontal='center')
HomNay['A2'].font = Font(name= 'Calibri', size=12)
HomNay['B2'] = 'Giờ lưu dữ liệu'
HomNay['B2'].alignment = Alignment(horizontal='center')
HomNay['B2'].font = Font(name= 'Calibri', size=12)
HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
HomNay['C2'].alignment = Alignment(horizontal='center')
HomNay['C2'].font = Font(name= 'Calibri', size=12)
HomNay['C3'] = 'Tổng số lượng sản xuất'
HomNay['D3'] = 'Cuộn cảm'
HomNay['E3'] = 'Cacbon tay chổi'
HomNay['F3'] = 'Hàn chổi'
HomNay['G3'] = 'Hàn chấu '
HomNay['H3'] = 'Đế vỏ nhỏ'
HomNay['I3'] = 'Tụ điện'
HomNay['J3'] = 'Cong chấu điện'
HomNay['K3'] = 'Bụi chì'
HomNay['L3'] = 'PP khác (nhiều hạng mục)'

for i in range(67,77):
    HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
    HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 
    
HomNay.column_dimensions['A'].width = 20
HomNay.column_dimensions['B'].width = 20
HomNay.column_dimensions['C'].width = 25
HomNay.column_dimensions['D'].width = 18
HomNay.column_dimensions['E'].width = 18
HomNay.column_dimensions['F'].width = 18
HomNay.column_dimensions['G'].width = 18
HomNay.column_dimensions['H'].width = 18
HomNay.column_dimensions['I'].width = 18
HomNay.column_dimensions['J'].width = 18
HomNay.column_dimensions['K'].width = 18
HomNay.column_dimensions['L'].width = 25

wb.remove(wb['Sheet'])

wb.save(f"excel/{mydate}_Ngay.xlsx")